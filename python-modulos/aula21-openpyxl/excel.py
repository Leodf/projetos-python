"""
https://openpyxl.readthedocs.io/en/stable/
pip install openpyxl
"""
import openpyxl
from random import uniform

"""
pedidos = openpyxl.load_workbook('./python-modulos/aula21-openpyxl/pedidos.xlsx')
nome_planilhas = pedidos.sheetnames
planilha1 = pedidos['Plan1']
# planilha1['b3'].value = 'Macarrão'

for linha in range(6, 16):
    numero_do_pedido = linha-1
    planilha1.cell(linha, 1).value = numero_do_pedido
    planilha1.cell(linha, 2).value = 'produto novo'

    qtde = round(uniform(1, 10))
    planilha1.cell(linha, 3).value = qtde

    preco = round(uniform(10, 100), 2)
    planilha1.cell(linha, 4).value = preco

    preco_total = preco*qtde
    planilha1.cell(linha, 5).value = preco_total

pedidos.save('./python-modulos/aula21-openpyxl/nova_planilha.xlsx')
"""
"""
for linha in planilha1:
    print(linha[0].value, linha[1].value, linha[2].value, linha[3].value, linha[4].value)
"""
planilha = openpyxl.Workbook()
planilha.create_sheet('Planilha1', 0)
planilha.create_sheet('Planilha2', 1)

planilha1 = planilha['Planilha1']
planilha2 = planilha['Planilha2']

for linha in range(1, 11):
    numero_do_pedido = linha-1
    planilha1.cell(linha, 1).value = numero_do_pedido
    planilha1.cell(linha, 2).value = 'produto novo'

    qtde = round(uniform(1, 10))
    planilha1.cell(linha, 3).value = qtde

    preco = round(uniform(10, 100), 2)
    planilha1.cell(linha, 4).value = preco

    preco_total = preco*qtde
    planilha1.cell(linha, 5).value = preco_total

for linha in range(1, 11):
    planilha2.cell(linha, 1).value = f'Leonardo {linha} {round(uniform(10, 100), 2)}'
    planilha2.cell(linha, 2).value = f'Maria {linha} {round(uniform(10, 100), 2)}'
    planilha2.cell(linha, 3).value = f'Joaozinho {linha} {round(uniform(10, 100), 2)}'

planilha.save('./python-modulos/aula21-openpyxl/nova_planilha.xlsx')